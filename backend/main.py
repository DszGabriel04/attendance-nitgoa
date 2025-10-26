
from fastapi import Body, FastAPI, Depends, HTTPException, Path, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse, RedirectResponse, HTMLResponse
from sqlalchemy import func, case, select

from sqlalchemy.orm import Session
import io
import secrets
import qrcode
from qrcode.constants import ERROR_CORRECT_M
import base64
import models, schemas
from typing import List, Dict, Set
from schemas import AttendanceRequest, QRCheck
from datetime import date
import threading
import time
from html_templates import ATTENDANCE_FORM_HTML, TOKEN_EXPIRED_HTML

# Add imports for Excel functionality
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

import redis
import os

REDIS_URL = os.getenv("REDIS_URL")                  #redis://red-d3tkrlqli9vc73bfkgi0:6379
r = redis.from_url(REDIS_URL, decode_responses=True)

app = FastAPI()

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "https://attendance-nitgoa.vercel.app",
        "https://attendance-nitgoa-jkm71scqp-russells-projects-bb2412f0.vercel.app",
        "http://localhost:8081",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/api/test_redis")
def test_redis():
    r.set("test_key", "ok", ex=10)
    return {"value": r.get("test_key")}


def get_db():                   # Dependency to get DB session
    db = models.SessionLocal()
    try:        yield db
    finally:    db.close()


@app.post("/faculty/login")
def faculty_login(email: str, password: str, db: Session = Depends(get_db)):
    faculty = db.query(models.Faculty).filter(models.Faculty.email == email).first()
    if not faculty:                        raise HTTPException(status_code=404, detail="Invalid email")
    if password != faculty.password_hash:  raise HTTPException(status_code=401, detail="Invalid password")      # plain password for now
    return {"message": "Login successful", "faculty_id": faculty.id} # need this


@app.post("/classes")
def create_class(class_data: schemas.StudentsBatchCreate, db: Session = Depends(get_db)):
    # check if faculty exists
    faculty = db.query(models.Faculty).filter(models.Faculty.id == class_data.faculty_id).first()
    if not faculty:
        raise HTTPException(status_code=404, detail=f"Faculty with id '{class_data.faculty_id}' does not exist")

    #create class info if it doesn't exist
    existing_class = db.query(models.Class).filter(models.Class.id == class_data.id).first()
    if existing_class:      raise HTTPException(status_code=400, detail=f"Class '{class_data.id}' already exists. Please delete the existing class before creating a new one.")
    db.add(models.Class(id=class_data.id, subject_name=class_data.subject_name,faculty_id=class_data.faculty_id))
    db.flush()

    #add students names with roll in the students database
    for student in class_data.students:
        existing_student = db.query(models.Student).filter(models.Student.id == student.id).first()
        if not existing_student:    db.add(models.Student(id=student.id, name=student.name))
    db.flush()

    # add student's attendance entries (initialized as absent)
    today = date.today()
    for student in class_data.students:
        existing_attendance = db.query(models.Attendance).filter(models.Attendance.class_id == class_data.id, models.Attendance.student_id == student.id, models.Attendance.date == today).first()
        if not existing_attendance:     db.add(models.Attendance(class_id=class_data.id, student_id=student.id, date=today, present=False))

    db.commit()
    return {"message": f"Class '{class_data.id}' created with {len(class_data.students)} students, attendance initialized for today"}


@app.delete("/classes/{class_id}")
def delete_class(class_id: str = Path(..., description="ID of the class to delete"), db: Session = Depends(get_db)):

    class_obj = db.query(models.Class).filter(models.Class.id == class_id).first()
    if not class_obj:   raise HTTPException(status_code=404, detail=f"Class with id '{class_id}' does not exist")

    db.query(models.Attendance).filter(models.Attendance.class_id == class_id).delete()
    db.delete(class_obj)
    db.commit()

    # delete students not belonging to any class from the students db
    all_students = db.query(models.Student).all()
    for student in all_students:
        has_attendance = db.query(models.Attendance).filter(models.Attendance.student_id == student.id).first()
        if not has_attendance:  db.delete(student)

    db.commit()
    return {"message": f"Class '{class_id}' and its attendance have been deleted. Students no longer in any class were also removed."}


@app.get("/classes")
def get_classes(db: Session = Depends(get_db)):
    today = date.today()

    # Count present attendance per class for today (only count if someone is marked present)
    # This ensures that attendance is only considered "taken" when at least one student is marked present,
    # not just when attendance records exist (which happens automatically on class creation with present=False)
    attendance_counts = ( select( models.Attendance.class_id, func.count(models.Attendance.id).label("count_present")).where(models.Attendance.date == today, models.Attendance.present == True).group_by(models.Attendance.class_id).subquery())

    # Join classes with attendance counts to determine if attendance has been actually taken today
    # "Yes" = at least one student marked present today, "No" = no students marked present today
    stmt = (select(models.Class.id, models.Class.subject_name, case( (attendance_counts.c.count_present > 0, "Yes"), else_="No").label("attendance_taken"))
        .outerjoin(attendance_counts, models.Class.id == attendance_counts.c.class_id))


    classes = db.execute(stmt).mappings().all()

    return ( [{"id": cls["id"], "subject_name": cls["subject_name"], "attendance_taken": cls["attendance_taken"]} for cls in classes])


@app.post("/classes/{class_id}/attendence")
def save_class_attendence(attendance_data: AttendanceRequest, class_id: str = Path(..., description="ID of the class"), db: Session = Depends(get_db)):
    attendance_date = date.today()

    class_obj = db.query(models.Class).filter(models.Class.id == class_id).first()
    if not class_obj:
        raise HTTPException(status_code=404, detail=f"Class '{class_id}' not found")

    created = 0
    errors: List[str] = []

    for item in attendance_data.attendees:
        student = db.query(models.Student).filter(models.Student.id == item.student_id).first()
        if not student:
            errors.append(f"Student '{item.student_id}' not found")
            continue

        # UPDATES ARE NOT ENABLED VIA POST; WILL RETURN ERROR
        existing = db.query(models.Attendance).filter(
            models.Attendance.class_id == class_id,
            models.Attendance.student_id == item.student_id,
            models.Attendance.date == attendance_date
        ).first()

        if existing:
            errors.append(f"Attendance already recorded for student '{item.student_id}'")
            continue

        new_rec = models.Attendance(
            class_id=class_id,
            student_id=item.student_id,
            date=attendance_date,
            present=item.present
        )
        db.add(new_rec)
        created += 1

    db.commit()

    return {
        "message": "Attendance saved",
        "class_id": class_id,
        "date": str(attendance_date),
        "created": created,
        "skipped": errors
    }


@app.put("/classes/{class_id}/attendance")
def update_attendance(attendance_data: AttendanceRequest, class_id: str = Path(..., description="ID of the class"), db: Session = Depends(get_db)):
    existing_rows = db.query(models.Attendance).filter(
        models.Attendance.class_id == class_id,
        models.Attendance.date == date.today()
    ).all()

    if not existing_rows:
        raise HTTPException(status_code=404, detail="No attendance records found for this class and date")

    existing_map = {r.student_id: r for r in existing_rows}

    updated = 0
    errors: List[str] = []

    for item in attendance_data.attendees:
        if item.student_id not in existing_map:
            errors.append(f"Student '{item.student_id}' has no existing record for this date")
            continue

        record = existing_map[item.student_id]
        if record.present != item.present:
            record.present = item.present
            updated += 1

    db.commit()

    return {
        "message": f"Updated {updated} record(s)",
        "errors": errors
    }


@app.get("/attendance/history/{class_id}") #     ----->  [] if course code doesnt exist else json of records
def get_attendance_history(class_id: str, db: Session = Depends(get_db)):
    class_obj = db.query(models.Class).filter(models.Class.id == class_id).first()
    if not class_obj:  return []    #raise HTTPException(status_code=404, detail="Class not found")

    # Fetch attendance records with student details
    attendance_records = (db.query( models.Attendance.date, models.Student.id.label("student_id"), models.Student.name.label("student_name"), models.Student.id, models.Attendance.present)
        .join(models.Student, models.Student.id == models.Attendance.student_id).filter(models.Attendance.class_id == class_obj.id) .order_by(models.Attendance.date.asc(), models.Student.id.asc()).all() )

    history = [ {"date": record.date, "student_id": record.student_id, "student_name": record.student_name, "status": "P" if record.present else "A"} for record in attendance_records]
    
    return {"class_code": class_id, "attendance_history": history}


@app.get("/attendance/export/{class_id}")
def export_attendance_excel(class_id: str, db: Session = Depends(get_db)):
    """
    Generate Excel file for attendance data of a specific class
    Returns Excel file as streaming response with formatted data
    """
    # Check if class exists
    class_obj = db.query(models.Class).filter(models.Class.id == class_id).first()
    if not class_obj:
        raise HTTPException(status_code=404, detail="Class not found")
    
    # Fetch attendance records with student details
    attendance_records = (
        db.query(
            models.Attendance.date,
            models.Student.id.label("student_id"),
            models.Student.name.label("student_name"),
            models.Attendance.present
        )
        .join(models.Student, models.Student.id == models.Attendance.student_id)
        .filter(models.Attendance.class_id == class_obj.id)
        .order_by(models.Attendance.date.asc(), models.Student.id.asc())
        .all()
    )
    
    if not attendance_records:
        raise HTTPException(status_code=404, detail="No attendance data found for this class")
    
    # Transform data for Excel format - organize by student and date
    dates = sorted(list(set([str(record.date) for record in attendance_records])))
    students = {}
    
    # Organize attendance data by student
    for record in attendance_records:
        student_id = record.student_id
        if student_id not in students:
            students[student_id] = {
                'name': record.student_name,
                'attendance': {}
            }
        students[student_id]['attendance'][str(record.date)] = 'P' if record.present else 'A'
    
    # Create Excel workbook with proper formatting
    wb = Workbook()
    ws = wb.active
    ws.title = f"Attendance - {class_id}"
    
    # Set up column headers (removed Total Present and Total Classes columns)
    headers = ['Roll No', 'Student Name'] + dates + ['Percentage']
    
    # Define Excel styling for professional appearance
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light red
    
    # Add column headers with styling (no title row)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add student data rows (starting from row 2 since no title row)
    row = 2
    for student_id in sorted(students.keys()):
        student_data = students[student_id]
        
        # Add basic student information
        ws.cell(row=row, column=1, value=student_id).border = border
        ws.cell(row=row, column=2, value=student_data['name']).border = border
        
        # Add attendance data with color coding
        present_count = 0
        total_classes = len(dates)
        
        for col, d8 in enumerate(dates, 3):
            status = student_data['attendance'].get(d8, '-')
            cell = ws.cell(row=row, column=col, value=status)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            # Apply color coding for attendance status
            if status == 'P':
                cell.fill = present_fill
                present_count += 1
            elif status == 'A':
                cell.fill = absent_fill
        
        # Calculate and add percentage (only percentage column now)
        percentage = (present_count / total_classes * 100) if total_classes > 0 else 0
        
        percentage_cell = ws.cell(row=row, column=len(dates) + 3, value=f"{percentage:.1f}%")
        percentage_cell.border = border
        percentage_cell.alignment = Alignment(horizontal='center')
        
        row += 1
    
    # Optimize column widths for better readability
    ws.column_dimensions['A'].width = 12  # Roll No
    ws.column_dimensions['B'].width = 25  # Student Name
    
    # Set width for date columns
    for col in range(3, len(dates) + 3):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 12
    
    # Set width for percentage column (only one summary column now)
    percentage_col_num = len(dates) + 3
    column_letter = get_column_letter(percentage_col_num)
    ws.column_dimensions[column_letter].width = 15
    
    # Save Excel file to memory buffer
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Generate filename with current date
    filename = f"attendance_{class_id}_{date.today().strftime('%Y%m%d')}.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    # Return Excel file as streaming response
    return StreamingResponse(
        io.BytesIO(excel_buffer.getvalue()),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers
    )



_tokens: dict[str, dict] = {}
_tokens_lock = threading.Lock()

# save_token, is_token_active and invalidate_token, helper functions to interact with the token dictionary in a safe way
def save_token(token: str, class_id: str) -> None:
    with _tokens_lock:
        _tokens[token] = {
            "active": True,
            "class_id": class_id,
            "submissions": set()
        }

def is_token_active(token: str) -> bool:
    with _tokens_lock:
        token_data = _tokens.get(token)
        return bool(token_data and token_data.get("active", False))

def invalidate_token(token: str) -> bool:
    with _tokens_lock:
        return _tokens.pop(token, None) is not None

def get_token_class_id(token: str) -> str | None:
    with _tokens_lock:
        token_data = _tokens.get(token)
        return token_data.get("class_id") if token_data else None

def get_token_submissions(token: str) -> set[str]:
    with _tokens_lock:
        token_data = _tokens.get(token)
        return token_data.get("submissions", set()).copy() if token_data else set()

def add_student_to_token(token: str, student_id: str) -> bool:
    with _tokens_lock:
        token_data = _tokens.get(token)
        if not token_data or not token_data.get("active", False):
            return False
        if student_id in token_data["submissions"]:
            return False
        token_data["submissions"].add(student_id)
        return True

def get_token_submission_stats(token: str) -> dict:
    with _tokens_lock:
        token_data = _tokens.get(token)
        submissions = token_data.get("submissions", set()) if token_data else set()
        return {
            "total_submissions": len(submissions),
            "all_students": list(submissions),
            "recent_submissions": len(submissions),
            "recent_students": list(submissions),
            "ready_for_bulk": len(submissions) > 0
        }

# builds the qrcode png image and returns it as bytes
def make_qr_png_bytes(data: str, box_size: int = 10, border: int = 4, error_correction=ERROR_CORRECT_M) -> bytes:
    qr = qrcode.QRCode(version=None, error_correction=error_correction, box_size=box_size, border=border)
    qr.add_data(data)
    qr.make(fit=True)

    qr_img = qr.make_image(fill_color="black", back_color="white")      # Pillow image
    buffer = io.BytesIO()
    qr_img.save(buffer, format="PNG") # type: ignore
    buffer.seek(0)
    return buffer.getvalue()        # returns a png image in bytes  


# 1) generated a qr code encodes /qr/validate?token=RANDOM_TOKEN; 
# 2) tokens generated randomly, stored in valid_tokens dict
# 3) on scanning redirects to above url
# 4) functin returns QR png as bytes + token, save token in frontend to invalidate later
# 5) base64 is one of the two modes of returning we chose; it return Json of token + data (qr code image that can be rendered using image tag)
@app.get("/qr/generate")
def generate_qr(
    request : Request,
    class_id: str = Query(..., description="Class ID for attendance"),
    length: int = Query(16, ge=1, le=256),
    box_size: int = Query(10, ge=1, le=40),
    border: int = Query(4, ge=0, le=20),
    as_base64: bool = Query(True),
    db: Session = Depends(get_db),
):
    # Verify class exists
    class_obj = db.query(models.Class).filter(models.Class.id == class_id).first()
    if not class_obj:
        raise HTTPException(status_code=404, detail=f"Class '{class_id}' not found")
    
    token = secrets.token_hex(length)
    save_token(token, class_id)

    # build absolute validation URL based on incoming request
    validation_url = str(request.url_for("validate_qr")) + f"?token={token}"
    # we build the QR code to automatically call the validate endpoint
    png_bytes = make_qr_png_bytes(data=validation_url, box_size=box_size, border=border)

    if as_base64:
        b64 = base64.b64encode(png_bytes).decode("ascii")
        return JSONResponse({"token": token, "data": f"data:image/png;base64,{b64}", "validation_url": validation_url})
    
    return StreamingResponse(io.BytesIO(png_bytes), media_type="image/png", headers={"X-QR-Token": token, "X-Validation-URL": validation_url})



# QR code validation page - shows form for student to enter roll number
@app.get("/qr/validate", name="validate_qr")
def validate_qr(token: str = Query(...), db: Session = Depends(get_db)):
    if not is_token_active(token):
        # Return expired token page
        return HTMLResponse(content=TOKEN_EXPIRED_HTML)
    
    # Render a minimal HTML page with embedded JS
    # JS sends device_id and session_id to /api/check_scan
    html = f"""
    <html>
      <body>
        <h3>Validating QR...</h3>
        <script>
          const session_id = "{token}";
          // Reuse or create a unique device_id
          const device_id = localStorage.getItem("device_id") || crypto.randomUUID();
          localStorage.setItem("device_id", device_id);

          fetch("/api/check_scan", {{
            method: "POST",
            headers: {{
              "Content-Type": "application/json"
            }},
            body: JSON.stringify({{ device_id, session_id }})
          }})
          .then(res => res.json())
          .then(data => {{
            if (data.allowed) {{
              window.location.href = "/qr/form?token=" + session_id;
            }} else {{
              document.body.innerHTML = "<h3>Already Scanned on this device!</h3>";
            }}
          }})
          .catch(() => {{
            document.body.innerHTML = "<h3>Server Error. Please try again.</h3>";
          }});
        </script>
      </body>
    </html>
    """
    return HTMLResponse(content=html)

# Get QR code status (number of students submitted) - optimized for bulk operations
@app.get("/qr/status")
def get_qr_status(token: str = Query(...), include_details: bool = Query(False)):
    if not is_token_active(token):
        raise HTTPException(status_code=410, detail="Token invalid or cancelled")
    
    class_id = get_token_class_id(token)
    
    if include_details:
        # Use detailed stats for bulk operations
        stats = get_token_submission_stats(token)
        response_data = {
            "submitted_count": stats.get("total_submissions", 0),
            "submitted_students": stats.get("all_students", []),
            "recent_submissions": stats.get("recent_submissions", 0),
            "recent_students": stats.get("recent_students", []),
            "class_id": class_id,
            "token_active": True,
            "ready_for_bulk_update": stats.get("ready_for_bulk", False),
            "last_updated": date.today().isoformat()
        }
    else:
        # Simple response for basic polling
        submitted_students = get_token_submissions(token)
        response_data = {
            "submitted_count": len(submitted_students),
            "submitted_students": list(submitted_students),
            "class_id": class_id,
            "token_active": True
        }
    
    return response_data

# Submit attendance via QR code
@app.post("/qr/submit-attendance")
def submit_attendance(payload: dict = Body(...), db: Session = Depends(get_db)):
    token = payload.get("token")
    student_id = payload.get("student_id")
    
    if not token or not student_id:
        raise HTTPException(status_code=400, detail="Token and student_id required")
    
    if not is_token_active(token):
        raise HTTPException(status_code=410, detail="Token invalid or cancelled")
    
    # Get class information
    class_id = get_token_class_id(token)
    if not class_id:
        raise HTTPException(status_code=400, detail="Invalid token configuration")
    
    # Check if student exists and is enrolled in this class
    # For now, we'll check if student exists in the students table
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        # Check if student has any attendance record for this class (indicating enrollment)
        has_record = db.query(models.Attendance).filter(
            models.Attendance.student_id == student_id,
            models.Attendance.class_id == class_id
        ).first()
        
        if not has_record:
            raise HTTPException(status_code=404, detail="Student not enrolled in this class")
    
    # Add student to token submissions
    if add_student_to_token(token, student_id):
        return {"message": "Attendance submission recorded", "student_id": student_id}
    else:
        # Check if student already submitted
        submitted_students = get_token_submissions(token)
        if student_id in submitted_students:
            raise HTTPException(status_code=409, detail="Attendance already submitted for this student")
        else:
            raise HTTPException(status_code=400, detail="Failed to record attendance")

# when the CANCEL button is clicked; it invalidates the current token associated with QR, 
# specifies that token saved earlier in dict, and commits attendance to database
@app.post("/qr/cancel")
def cancel_qr(payload: dict = Body(...), db: Session = Depends(get_db)):
    token = payload.get("token")
    if not token:
        raise HTTPException(status_code=400, detail="token required")
    
    if not is_token_active(token):
        raise HTTPException(status_code=404, detail="Token not found")
    
    # Get submitted students and class info
    submitted_students = get_token_submissions(token)
    class_id = get_token_class_id(token)
    
    attendance_date = date.today()
    marked_present = 0
    errors = []
    
    if class_id and submitted_students:
        # Bulk update approach - prepare all operations first, then execute
        try:
            # Get all existing attendance records for this class and date in a single query
            existing_records = db.query(models.Attendance).filter(
                models.Attendance.class_id == class_id,
                models.Attendance.date == attendance_date,
                models.Attendance.student_id.in_(submitted_students)
            ).all()
            
            # Create a map of existing records for quick lookup
            existing_map = {record.student_id: record for record in existing_records}
            
            # Process all submitted students
            records_to_update = []
            records_to_create = []
            
            for student_id in submitted_students:
                if student_id in existing_map:
                    # Update existing record
                    existing_record = existing_map[student_id]
                    if not existing_record.present:  # Only update if not already present
                        existing_record.present = True
                        records_to_update.append(student_id)
                        marked_present += 1
                else:
                    # Create new record
                    new_record = models.Attendance(
                        class_id=class_id,
                        student_id=student_id,
                        date=attendance_date,
                        present=True
                    )
                    db.add(new_record)
                    records_to_create.append(student_id)
                    marked_present += 1
            
            # Commit all changes in a single transaction
            db.commit()
            
            # Log the bulk operation for debugging
            if records_to_update:
                print(f"Bulk updated {len(records_to_update)} existing records: {records_to_update}")
            if records_to_create:
                print(f"Bulk created {len(records_to_create)} new records: {records_to_create}")
                
        except Exception as e:
            db.rollback()
            errors.append(f"Bulk update failed: {str(e)}")
            print(f"Bulk update error: {str(e)}")
    
    # #### Clear Redis cache for this session ####
    keys = r.keys(f"{token}:*")
    if keys:
        r.delete(*keys)

    # Invalidate token
    invalidate_token(token)
    
    return {
        "token": token, 
        "cancelled": True,
        "students_marked_present": marked_present,
        "submitted_students": list(submitted_students),
        "errors": errors
    }


@app.post("/api/check_scan")
def check_scan(data: QRCheck):
    key = f"{data.session_id}:{data.device_id}"
    if r.exists(key):
        return {"allowed": False}
    else:
        r.set(key, "1")
        return {"allowed": True}